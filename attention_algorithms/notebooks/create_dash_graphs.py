import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
import os

if __name__ == "__main__":
    sns.set_theme()
    cwd = os.getcwd().split(os.path.sep)
    while cwd[-1] != "stage_4_gm":
        os.chdir("..")
        cwd = os.getcwd().split(os.path.sep)

    # the file
    dir = os.path.join(os.getcwd(), ".cache", "plots")
    wb = openpyxl.load_workbook(os.path.join(dir, "dash_board.xlsx"))

    # graphs parameters
    ylims = [(0.15, 0.4), (0.2, 0.65), (0.1, 0.6), (0.5, 0.75), (0.25, 0.5)]
    tl = ["F1", "PR", "RC", "ROC", "AUPRC"]

    # cells for the entropia
    reg_mul_s = [0, 0.001, 0.002, 0.003, 0.004, 0.005, 0.08, 0.1, 0.4]

    # for the entropia
    col = ["A", "B", "C", "D", "E", "G"]

    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################

    ####################################
    # the line plots for the reg study #
    ####################################

    rows = [7, 8, 9, 10, 11, 12, 13]
    fig, axes = plt.subplots(5, 2, figsize=(20, 20))
    for r in reg_mul_s:
        ws = wb["reg_mul=" + str(r)]
        temp = np.zeros((len(rows), len(col)))
        for i in range(len(rows)):
            for j in range(len(col)):
                temp[i, j] = ws[col[j] + str(rows[i])].value

        k = 1
        for i in range(axes.shape[0]):
            ax = axes[i, 0]
            if k <= 5:
                ax.plot(temp[:, 0], temp[:, k], label="reg_mul=" + str(r))
                ax.set_ylim(ylims[k - 1][0], ylims[k - 1][1])
                ax.set_title(tl[k - 1] + " -- sum_agreg")
                ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop={"size": 6})
                k += 1

    # for the CLS map
    rows = [19, 20, 21, 22, 23, 24, 25]
    for r in reg_mul_s:
        ws = wb["reg_mul=" + str(r)]
        temp = np.zeros((len(rows), len(col)))
        for i in range(len(rows)):
            for j in range(len(col)):
                temp[i, j] = ws[col[j] + str(rows[i])].value

        k = 1
        for i in range(axes.shape[0]):
            ax = axes[i, 1]
            if k <= 5:
                ax.plot(temp[:, 0], temp[:, k], label="reg_mul=" + str(r))
                ax.set_title(tl[k - 1] + " -- cls")
                ax.set_ylim(ylims[k - 1][0], ylims[k - 1][1])
                ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop={"size": 6})
                k += 1

    plt.subplots_adjust(left=0.1,
                        bottom=0.1,
                        right=0.9,
                        top=0.9,
                        wspace=0.4,
                        hspace=0.4)
    plt.savefig(os.path.join(os.getcwd(), "buff.png"))

    ws = None
    if "metric graphs" in wb.sheetnames:
        std = wb.get_sheet_by_name('metric graphs')
        wb.remove_sheet(std)

    ws = wb.create_sheet("metric graphs")
    ws = wb["metric graphs"]

    img = openpyxl.drawing.image.Image(os.path.join(os.getcwd(), "buff.png"))
    # img.anchor(ws.cell('A1'))
    ws.add_image(img, 'A1')


    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################

    ###############################################
    # the barplots for the entire dataset metrics #
    ###############################################

    cols = ["A", "B", "C", "D", "F"]
    met = ["F1", "P", "R", "ROC", "AUPRC"]
    ylims = [(0.2, 0.45), (0.2, 0.65), (0.2, 1), (0.5, 0.8), (0.25, 0.5)]
    fig, axes = plt.subplots(5, 3, figsize=(30, 20))

    for i in range(axes.shape[0]):
        ax0 = axes[i, 0]  # ent
        ax1 = axes[i, 1]  # cls
        ax2 = axes[i, 2]

        ent = []
        cls = []
        flow = []
        for r in reg_mul_s:
            ws = wb["reg_mul=" + str(r)]
            cls.append(ws[cols[i] + str(38)].value)
            ent.append(ws[cols[i] + str(33)].value)

        ws = wb["flow_study"]
        flow.append(ws[cols[i] + str(3)].value)
        flow.append(ws[cols[i] + str(6)].value)

        ax0.bar(list(range(len(ent))), ent)
        ax0.set_xticks(list(range(len(ent))))
        ax0.set_xticklabels(["r="+str(x) for x in reg_mul_s], fontsize=8, rotation=45)
        ax0.set_title(met[i] + " -- sum_agreg")
        ax0.set_ylim(ylims[i][0], ylims[i][1])


        ax1.bar(list(range(len(cls))), cls)
        ax1.set_xticks(list(range(len(cls))))
        ax1.set_xticklabels(["r=" + str(x) for x in reg_mul_s], fontsize=8, rotation=60)
        ax1.set_title(met[i] + " -- cls")
        ax1.set_ylim(ylims[i][0], ylims[i][1])

        ax2.bar(list(range(len(flow))), flow)
        ax2.set_xticks(list(range(len(flow))))
        ax2.set_xticklabels(["avg_agreg", "max_agreg"], fontsize=8, rotation=60)
        ax2.set_title(met[i] + " -- flow")
        ax2.set_ylim(ylims[i][0], ylims[i][1])

    plt.subplots_adjust(left=0.1,
                        bottom=0.1,
                        right=0.9,
                        top=0.9,
                        wspace=0.4,
                        hspace=0.4)

    plt.savefig(os.path.join(os.getcwd(), "buff2.png"))

    ws = None
    if "all DS bar plots" in wb.sheetnames:
        std = wb.get_sheet_by_name('all DS bar plots')
        wb.remove_sheet(std)

    ws = wb.create_sheet("all DS bar plots")
    ws = wb["all DS bar plots"]

    img = openpyxl.drawing.image.Image(os.path.join(os.getcwd(), "buff2.png"))
    # img.anchor(ws.cell('A1'))
    ws.add_image(img, 'A1')


    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################

    #######################################
    # add images only for the reg_mul = 0 #
    #######################################
    ylims = [(0.15, 0.4), (0.2, 0.65), (0.1, 0.6), (0.5, 0.75), (0.25, 0.5)]
    rows = [7, 8, 9, 10, 11, 12, 13]
    fig, axes = plt.subplots(5, 2, figsize=(20, 20))
    for r in [0]:
        ws = wb["reg_mul=" + str(r)]
        temp = np.zeros((len(rows), len(col)))
        for i in range(len(rows)):
            for j in range(len(col)):
                temp[i, j] = ws[col[j] + str(rows[i])].value

        k = 1
        for i in range(axes.shape[0]):
            ax = axes[i, 0]
            if k <= 5:
                ax.plot(temp[:, 0], temp[:, k], label="reg_mul=" + str(r))
                ax.set_ylim(ylims[k - 1][0], ylims[k - 1][1])
                ax.set_title(tl[k - 1] + " -- sum_agreg")
                ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop={"size": 6})
                k += 1



    rows = [19, 20, 21, 22, 23, 24, 25]
    for r in [0]:
        ws = wb["reg_mul=" + str(r)]
        temp = np.zeros((len(rows), len(col)))
        for i in range(len(rows)):
            for j in range(len(col)):
                temp[i, j] = ws[col[j] + str(rows[i])].value

        k = 1
        for i in range(axes.shape[0]):
            ax = axes[i, 1]
            if k <= 5:
                ax.plot(temp[:, 0], temp[:, k], label="reg_mul=" + str(r))
                ax.set_title(tl[k - 1] + " -- cls")
                ax.set_ylim(ylims[k - 1][0], ylims[k - 1][1])
                ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), prop={"size": 6})
                k += 1

    plt.subplots_adjust(left=0.1,
                        bottom=0.1,
                        right=0.9,
                        top=0.9,
                        wspace=0.4,
                        hspace=0.4)
    plt.savefig(os.path.join(os.getcwd(), "buff3.png"))

    ws = None
    if "heads selection criterion" in wb.sheetnames:
        std = wb.get_sheet_by_name('heads selection criterion')
        wb.remove_sheet(std)

    ws = wb.create_sheet("heads selection criterion")
    ws = wb["heads selection criterion"]

    img = openpyxl.drawing.image.Image(os.path.join(os.getcwd(), "buff3.png"))
    # img.anchor(ws.cell('A1'))
    ws.add_image(img, 'A1')



    wb.save(os.path.join(dir, "dash_board.xlsx"))
    wb.close()

    os.remove("buff.png")
    os.remove("buff2.png")
    os.remove("buff3.png")
